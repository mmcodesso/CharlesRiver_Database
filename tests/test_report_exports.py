from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

from generator_dataset.reports import load_report_catalog
from generator_dataset.settings import load_settings


EXPECTED_REPORT_SLUGS = [
    "monthly-income-statement",
    "monthly-balance-sheet",
    "monthly-indirect-cash-flow",
    "monthly-revenue-and-gross-margin",
    "ar-aging",
    "ap-aging",
    "budget-vs-actual-by-cost-center",
    "sales-and-margin-by-collection-and-style",
    "monthly-work-center-utilization",
    "headcount-by-cost-center-and-job-family",
    "approval-and-sod-review",
    "potential-anomaly-review",
]


def _header_row(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    try:
        return [str(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    finally:
        workbook.close()


def test_report_catalog_contains_curated_v1_pack() -> None:
    catalog = load_report_catalog()
    assert [report.slug for report in catalog] == EXPECTED_REPORT_SLUGS


def test_report_exports_create_curated_artifacts(
    report_validation_dataset_artifacts: dict[str, object],
) -> None:
    report_root = Path(report_validation_dataset_artifacts["report_output_dir"])

    for report in load_report_catalog():
        asset_dir = report_root / report.area / report.process_group / report.slug
        excel_path = asset_dir / f"{report.slug}.xlsx"
        csv_path = asset_dir / f"{report.slug}.csv"
        preview_path = asset_dir / "preview.json"

        assert asset_dir.exists(), f"Missing asset directory for {report.slug}"
        assert excel_path.exists() and excel_path.stat().st_size > 0
        assert csv_path.exists() and csv_path.stat().st_size > 0
        assert preview_path.exists() and preview_path.stat().st_size > 0


def test_report_exports_preserve_columns_and_preview_schema(
    report_validation_dataset_artifacts: dict[str, object],
) -> None:
    context = report_validation_dataset_artifacts["context"]
    report_root = Path(report_validation_dataset_artifacts["report_output_dir"])

    for report in load_report_catalog():
        asset_dir = report_root / report.area / report.process_group / report.slug
        csv_frame = pd.read_csv(asset_dir / f"{report.slug}.csv")
        preview = json.loads((asset_dir / "preview.json").read_text(encoding="utf-8"))
        workbook_headers = _header_row(asset_dir / f"{report.slug}.xlsx")
        effective_preview_limit = min(report.preview_row_limit, context.settings.report_preview_row_count)

        assert not csv_frame.empty, f"Expected rows in exported report {report.slug}"
        assert preview["slug"] == report.slug
        assert preview["title"] == report.title
        assert preview["area"] == report.area
        assert preview["processGroup"] == report.process_group
        assert preview["rowCount"] == len(csv_frame.index)
        assert preview["previewRowLimit"] == effective_preview_limit
        assert preview["previewRowCount"] == min(len(csv_frame.index), effective_preview_limit)
        assert preview["columns"] == csv_frame.columns.tolist()
        assert preview["columns"] == workbook_headers
        assert len(preview["rows"]) == preview["previewRowCount"]
        assert preview["generatedAt"]


def test_report_manifest_and_docs_include_curated_paths() -> None:
    manifest_text = Path("src/generated/reportManifest.js").read_text(encoding="utf-8")
    collections_text = Path("src/generated/reportDocCollections.js").read_text(encoding="utf-8")
    sidebar_text = Path("sidebars.js").read_text(encoding="utf-8")
    analytics_hub = Path("docs/analytics/index.md").read_text(encoding="utf-8")
    reports_hub = Path("docs/analytics/reports/index.md").read_text(encoding="utf-8")
    financial_reports = Path("docs/analytics/reports/financial.md").read_text(encoding="utf-8")
    managerial_reports = Path("docs/analytics/reports/managerial.md").read_text(encoding="utf-8")
    audit_reports = Path("docs/analytics/reports/audit.md").read_text(encoding="utf-8")
    docusaurus_config = Path("docusaurus.config.js").read_text(encoding="utf-8")

    for path in [
        Path("docs/analytics/reports/index.md"),
        Path("docs/analytics/reports/financial.md"),
        Path("docs/analytics/reports/managerial.md"),
        Path("docs/analytics/reports/audit.md"),
    ]:
        assert path.exists(), f"Missing report doc: {path}"

    assert '"analytics/reports/index"' in sidebar_text
    assert '"analytics/reports/financial"' in sidebar_text
    assert '"analytics/reports/managerial"' in sidebar_text
    assert '"analytics/reports/audit"' in sidebar_text
    assert '"outputs/site"' in docusaurus_config
    assert "Reports Hub" in analytics_hub
    assert "Download Excel" in reports_hub
    assert "reportAreaCollections.financial" in financial_reports
    assert "reportAreaCollections.managerial" in managerial_reports
    assert "reportAreaCollections.audit" in audit_reports

    for report in load_report_catalog():
        asset_base_path = f"/reports/{report.area}/{report.process_group}/{report.slug}"
        assert report.slug in manifest_text
        assert report.slug in collections_text
        assert f'"previewPath": "{asset_base_path}/preview.json"' in manifest_text
        assert f'"excelPath": "{asset_base_path}/{report.slug}.xlsx"' in manifest_text
        assert f'"csvPath": "{asset_base_path}/{report.slug}.csv"' in manifest_text


def test_report_settings_require_sqlite_export(tmp_path: Path) -> None:
    config_path = tmp_path / "settings.yaml"
    config_path.write_text(
        yaml.safe_dump(
            {
                "random_seed": 20260401,
                "fiscal_year_start": "2026-01-01",
                "fiscal_year_end": "2026-12-31",
                "company_name": "Charles River Home Furnishings, Inc.",
                "short_name": "CharlesRiver",
                "base_url": "https://charlesriver.accountinganalyticshub.com",
                "tax_rate": 0.065,
                "employee_count": 48,
                "customer_count": 60,
                "supplier_count": 35,
                "item_count": 90,
                "warehouse_count": 2,
                "export_sqlite": False,
                "export_excel": False,
                "export_support_excel": False,
                "export_csv_zip": False,
                "export_reports": True,
                "anomaly_mode": "none",
                "sqlite_path": "outputs/{short_name}_validation.sqlite",
                "excel_path": "outputs/{short_name}_validation.xlsx",
                "support_excel_path": "outputs/{short_name}_validation_support.xlsx",
                "csv_zip_path": "outputs/{short_name}_validation_csv.zip",
                "report_output_dir": "outputs/site/reports",
                "report_preview_row_count": 25,
                "generation_log_path": "outputs/generation_validation.log",
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="export_reports requires export_sqlite"):
        load_settings(config_path)
